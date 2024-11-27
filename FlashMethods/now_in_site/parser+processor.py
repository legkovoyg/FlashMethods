import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, PrivateAttr
from abc import ABC, abstractmethod
import numpy as np
from typing import List, Optional, Dict

import equation as eqs


# === Абстрактный парсер ===
class Parser(ABC):
    @abstractmethod
    def load_data(self, file_path: str):
        pass


# === Парсер данных из Excel ===
class ExcelDataParser(Parser):
    def load_data(self, file_path: str) -> Dict:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            data = {}
            # Итерация по колонкам
            for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
                header = ws.cell(row=1, column=col_idx).value
                if header is not None:
                    # Фильтрация данных, исключая None и заголовок
                    col_data = [cell for cell in col[1:] if cell is not None]
                    data[header] = col_data
            return data
        except InvalidFileException:
            print("ExcelDataParser: ошибка при открытии файла")
            return None


# === Парсер матрицы BIPs из Excel ===
class ExcelBIPSParser(Parser):
    def load_data(self, file_path: str) -> pd.DataFrame:
        try:
            BIPS = pd.read_excel(file_path, header=None)
            return BIPS
        except InvalidFileException:
            print("ExcelBIPSParser: ошибка при открытии файла")
            return None


class JSONDataParser(Parser):
    def load_data(self, file_path: str) -> Dict:
        try:
            with open(file_path, "r") as file:
                return json.load(file)
        except json.JSONDecodeError:
            print("JSONDataParser: ошибка при чтении JSON файла")
            return None


# === DTO для компонента жидкости ===
class FluidComponentDTO(BaseModel):
    """
    Каждый компонент описывается по вот этой штуке
    По идее не все обязательные - например cpen не нужен если используется НЕ peneloux версия уравнений
    T_boil и density_liq_phase могут не использоваться (вроде) если не считается энтальпия, Cp и Cv
    """

    component_name: str
    z: float
    mass: float
    Pkr: float
    Tkr: float
    Vkr: float
    w: float
    cpen: float
    T_boil: float
    density_liq_phase: float


# === DTO для жидкости ===
class FluidDTO(BaseModel):
    """
    Каюсь использовал typing но в Pydantic чето не пошло тут, потом могу переделать
    BIPS могут быть опциональны, из-за них меняется форма фигуры (см. graphs)
    """

    components: List[FluidComponentDTO]
    BIPs: Optional[pd.DataFrame] = None
    _N: int = PrivateAttr(default=0)  # Приватное поле для количества компонентов
    _c: np.ndarray = PrivateAttr(
        default_factory=lambda: np.zeros((0, 0), dtype=np.float64)
    )  # Приватная матрица BIPs

    class Config:
        arbitrary_types_allowed = True  # Позволяет использовать DataFrame и ndarray

    def __init__(self, **data):
        super().__init__(**data)
        # Установим количество компонентов
        self._N = len(self.components)
        # Преобразуем DataFrame в numpy-матрицу или инициализируем нулевую
        if self.BIPs is not None and not self.BIPs.empty:
            self._c = self.BIPs.to_numpy()
        else:
            self._c = np.zeros((self._N, self._N), dtype=np.float64)


# === Маппер для данных ===
class FluidMapper:
    @staticmethod
    def map_to_fluid_dto(data: Dict, bips: Optional[pd.DataFrame] = None) -> FluidDTO:
        components = [
            FluidComponentDTO(
                component_name=data["component_name"][i],
                z=float(data["z"][i]),
                mass=float(data["mass"][i]),
                Pkr=float(data["Pkr"][i]),
                Tkr=float(data["Tkr"][i]),
                Vkr=float(data["Vkr"][i]),
                w=float(data["w"][i]),
                cpen=float(data["cpen"][i]),
                T_boil=float(data["T_boil"][i]),
                density_liq_phase=float(data["density_liq_phase"][i]),
            )
            for i in range(len(data["component_name"]))
        ]
        return FluidDTO(components=components, BIPs=bips)


# === Универсальный обработчик данных ===
class FluidDataProcessor:
    def __init__(self, parser: Parser):
        self.parser = parser

    def process(self, file_path: str, bips_path: Optional[str] = None) -> FluidDTO:
        # Загрузка данных
        data = self.parser.load_data(file_path)
        if not data:
            raise ValueError("Не удалось загрузить данные.")

        # Загрузка BIPs, если указан путь
        bips = None
        if bips_path:
            bips_parser = ExcelBIPSParser()
            bips = bips_parser.load_data(bips_path)

        # Маппинг данных в FluidDTO
        return FluidMapper.map_to_fluid_dto(data, bips)


# === Основной блок ===
if __name__ == "__main__":
    # обработчик
    excel_parser = ExcelDataParser()
    processor = FluidDataProcessor(parser=excel_parser)
    # Присваиваем в переменную объект dto
    fluid = processor.process(
        file_path="data/input_data.xlsx", bips_path="data/BIPSS.xlsx"
    )
    # Кидаем в нужный класс объект dto (пользователь сам выбирает уравнение состояния)
    calculate_PR = eqs.PR_Flash(fluid)
    # Такой вот результат расчета
    (
        PR_W,  # Коэффициент разделения фаз (доля газовой фазы)
        PR_Z_v,  # Фактор сжимаемости газовой фазы
        PR_Z_l,  # Фактор сжимаемости жидкой фазы
        PR_x_i,  # Мольные доли компонентов в жидкой фазе
        PR_y_i,  # Мольные доли компонентов в газовой фазе
        PR_Stable,  # Флаг стабильности (1 - стабильно, 0 - нестабильно)
        PR_m,  # Количество итераций, выполненных при расчетах
        PR_enthalpy,  # Энтальпия смеси
        PR_enthalpy_w,  # Энтальпия газовой фазы
        PR_enthalpy_l,  # Энтальпия жидкой фазы
        PR_Cp,  # Теплоемкость смеси при постоянном давлении
        PR_Cp_w,  # Теплоемкость газовой фазы при постоянном давлении
        PR_Cp_l,  # Теплоемкость жидкой фазы при постоянном давлении
        PR_Cv,  # Теплоемкость смеси при постоянном объеме
        PR_Cv_w,  # Теплоемкость газовой фазы при постоянном объеме
        PR_Cv_l,  # Теплоемкость жидкой фазы при постоянном объеме
        PR_volume,  # Удельный объем смеси
        PR_VolumeMy_y,  # Удельный объем газовой фазы
        PR_VolumeMy_x,  # Удельный объем жидкой фазы
        PR_density,  # Плотность смеси
        PR_density_y,  # Плотность газовой фазы
        PR_density_x,  # Плотность жидкой фазы
    ) = calculate_PR.vle(10, 300)

    """
    С PR_peneloux пока проблема, его лучше не трогать
    """
    # calculate_PR_peneloux = eqs.PR_peneloux_Flash(fluid)
    # result_PR_peneloux = calculate_PR_peneloux.vle(10, 300)

    calculate_SRK = eqs.SRK_Flash(fluid)
    (
        SRK_W,
        SRK_Z_v,
        SRK_Z_l,
        SRK_x_i,
        SRK_y_i,
        SRK_Stable,
        SRK_m,
        SRK_enthalpy,
        SRK_enthalpy_w,
        SRK_enthalpy_l,
        SRK_Cp,
        SRK_Cp_w,
        SRK_Cp_l,
        SRK_Cv,
        SRK_Cv_w,
        SRK_Cv_l,
        SRK_volume,
        SRK_VolumeMy_y,
        SRK_VolumeMy_x,
        SRK_density,
        SRK_density_y,
        SRK_density_x,
    ) = calculate_SRK.vle(10, 300)

    calculate_SRK_Peneloux = eqs.SRK_peneloux_Flash(fluid)
    (
        SRK_peneloux_W,
        SRK_peneloux_Z_v,
        SRK_peneloux_Z_l,
        SRK_peneloux_x_i,
        SRK_peneloux_y_i,
        SRK_peneloux_Stable,
        SRK_peneloux_m,
        SRK_peneloux_enthalpy,
        SRK_peneloux_enthalpy_w,
        SRK_peneloux_enthalpy_l,
        SRK_peneloux_Cp,
        SRK_peneloux_Cp_w,
        SRK_peneloux_Cp_l,
        SRK_peneloux_Cv,
        SRK_peneloux_Cv_w,
        SRK_peneloux_Cv_l,
        SRK_peneloux_volume,
        SRK_peneloux_VolumeMy_y,
        SRK_peneloux_VolumeMy_x,
        SRK_peneloux_density,
        SRK_peneloux_density_y,
        SRK_peneloux_density_x,
    ) = calculate_SRK.vle(10, 300)

    print(
        f"Доля пара в смеси по разным уравениням состояния равна {SRK_W}, {PR_W}, {SRK_peneloux_W}"
    )

"""
Примечание

1
Пока что хз какое сделать dto для совершенного расчета (не садился за это)
С одной стороны - это свойство флюида, с другой стороны это флюид при определенном P, T

2 
Проверка стабильности зашита уже в расчетки - см. # Liquid phase calculation и # Поиск жидкой фазы
"""
