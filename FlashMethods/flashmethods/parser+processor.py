# === Импорты ===
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from abc import ABC, abstractmethod
from typing import List, Optional, Dict
import re
import numpy as np
import logging

# === Настройка Логирования ===
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# === DTO Классы ===
class FluidComponentDTO:
    def __init__(self, component_name: str, z: float, mass: float, Pkr: float, Tkr: float,
                 Vkr: float, w: float, cpen: float, T_boil: float, density_liq_phase: float):
        self.component_name = component_name
        self.z = z
        self.mass = mass
        self.Pkr = Pkr
        self.Tkr = Tkr
        self.Vkr = Vkr
        self.w = w
        self.cpen = cpen
        self.T_boil = T_boil
        self.density_liq_phase = density_liq_phase

    def __repr__(self):
        return (f"FluidComponentDTO(component_name={self.component_name}, z={self.z}, mass={self.mass}, "
                f"Pkr={self.Pkr}, Tkr={self.Tkr}, Vkr={self.Vkr}, w={self.w}, "
                f"cpen={self.cpen}, T_boil={self.T_boil}, density_liq_phase={self.density_liq_phase})")


class FluidDTO:
    def __init__(self, components: List[FluidComponentDTO], BIPs: Optional[pd.DataFrame], p: float, t: float):
        self.components = components
        self.BIPs = BIPs
        self.p = p
        self.t = t

# === Класс Flash ===
class SRK_peneloux_Flash:
    def __init__(self, fluid: FluidDTO):
        self.fluid = fluid

    def calculate(self):
        # Простая демонстрация, в реальности здесь будет сложная логика расчета
        return self.fluid.components

# === Абстрактный парсер ===
class Parser(ABC):
    @abstractmethod
    def load_data(self, file_path: str) -> Optional[Dict]:
        pass

# === Парсеры Excel ===
class ExcelDataParser:
    def load_data(self, file_path: str) -> Optional[Dict]:
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            data = {}
            # Итерация по колонкам
            for col_idx, col in enumerate(ws.iter_cols(values_only=True), start=1):
                header = ws.cell(row=1, column=col_idx).value
                if header is not None:
                    # Фильтрация данных, исключая None и заголовок
                    col_data = [cell for cell in col[1:] if cell is not None]
                    data[header] = col_data
            logger.info(f"ExcelDataParser: успешно загружены данные из {file_path}")
            return data
        except InvalidFileException:
            logger.error("ExcelDataParser: ошибка при открытии файла.")
            return None
        except Exception as e:
            logger.error(f"ExcelDataParser: произошла ошибка - {e}")
            return None

class ExcelBIPSParser:
    def load_data(self, file_path: str) -> Optional[pd.DataFrame]:
        try:
            BIPS = pd.read_excel(file_path, header=None)
            logger.info(f"ExcelBIPSParser: успешно загружены BIPs из {file_path}")
            return BIPS
        except InvalidFileException:
            logger.error("ExcelBIPSParser: ошибка при открытии файла.")
            return None
        except Exception as e:
            logger.error(f"ExcelBIPSParser: произошла ошибка - {e}")
            return None

class ExcelParser(Parser):
    def __init__(self):
        self.data_parser = ExcelDataParser()
        self.bips_parser = ExcelBIPSParser()

    def load_data(self, file_path: str, bips_path: Optional[str] = None) -> Optional[Dict]:
        try:
            data = self.data_parser.load_data(file_path)
            if data is None:
                logger.error("ExcelParser: Не удалось загрузить данные из Excel.")
                return None

            BIPS = None
            if bips_path:
                BIPS = self.bips_parser.load_data(bips_path)
                if BIPS is None:
                    logger.error("ExcelParser: Не удалось загрузить данные BIPs из Excel.")
                    return None

            data['BIPs'] = BIPS
            return data
        except Exception as e:
            logger.error(f"ExcelParser: произошла ошибка - {e}")
            return None

# === Абстрактный Экстрактор для CHC ===
class CHCExtractor(ABC):
    def __init__(self, content: str):
        self.content = content

    @abstractmethod
    def extract(self):
        pass

# === Конкретные Экстракторы CHC ===
class ShortCompNameExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[str]]:
        match = re.search(r"<Short Comp Name>\s*([\s\S]*?)\n(?=<|$)", self.content)
        if match:
            raw_list = re.findall(r'"([^"]+)"', match.group(1))
            component_names = [item.strip() for item in raw_list]
            logger.info("ShortCompNameExtractor: успешно извлечены component_name")
            return {'component_name': component_names}
        else:
            logger.warning("Тег <Short Comp Name> не найден в файле.")
            return {'component_name': []}

class CompAmountExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Comp Amount>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            z = [float(value) for value in raw_values][2:]
            logger.info("CompAmountExtractor: успешно извлечены z")
            return {'z': z}
        else:
            logger.warning("Тег <Comp Amount> не найден в файле.")
            return {'z': []}

class MolecularWeightsExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Mwn>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            mass = [float(value) for value in raw_values][2:]
            logger.info("MolecularWeightsExtractor: успешно извлечены mass")
            return {'mass': mass}
        else:
            logger.warning("Тег <Mwn> не найден в файле.")
            return {'mass': []}

class CriticalPressureExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Pc>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            pressures_atm = [float(value) for value in raw_values][2:]
            pressures_mpa = [value * 0.101325 for value in pressures_atm]
            logger.info("CriticalPressureExtractor: успешно извлечены Pkr")
            return {'Pkr': pressures_mpa}
        else:
            logger.warning("Тег <Pc> не найден в файле.")
            return {'Pkr': []}

class CriticalTemperatureExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Tc>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            temperatures = [float(value) for value in raw_values][2:]
            logger.info("CriticalTemperatureExtractor: успешно извлечены Tkr")
            return {'Tkr': temperatures}
        else:
            logger.warning("Тег <Tc> не найден в файле.")
            return {'Tkr': []}

class CriticalVolumeExtractor(CHCExtractor):
    R = 0.082057  # Газовая постоянная в L·atm/(K·mol)

    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Vc>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            vc_over_r = [float(value) for value in raw_values][2:]
            vc = [value * self.R * 1000 for value in vc_over_r]
            logger.info("CriticalVolumeExtractor: успешно извлечены Vkr")
            return {'Vkr': vc}
        else:
            logger.warning("Тег <Vc> не найден в файле.")
            return {'Vkr': []}

class AcentricFactorExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Acentric Factor>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            w = [float(value) for value in raw_values][2:]
            logger.info("AcentricFactorExtractor: успешно извлечены w")
            return {'w': w}
        else:
            logger.warning("Тег <Acentric Factor> не найден в файле.")
            return {'w': []}

class BoilingTemperatureExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Tb>[\s\S]*?([\d.,\s]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"[\d.]+", match.group(1))
            T_boil = [float(value) for value in raw_values][2:]
            logger.info("BoilingTemperatureExtractor: успешно извлечены T_boil")
            return {'T_boil': T_boil}
        else:
            logger.warning("Тег <Tb> не найден в файле.")
            return {'T_boil': []}

class LiquidDensityExtractor(CHCExtractor):
    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Density>[\s\S]*?([\d.eE,+\s-]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"-?[\d.eE+-]+", match.group(1))
            density = [float(value) if float(value) > 0 else 0 for value in raw_values][2:]
            logger.info("LiquidDensityExtractor: успешно извлечены density_liq_phase")
            return {'density_liq_phase': density}
        else:
            logger.warning("Тег <Density> не найден в файле.")
            return {'density_liq_phase': []}

class CpenSRKExtractor(CHCExtractor):
    R = 0.082057  # Газовая постоянная в L·atm/(K·mol)

    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Cpen SRK>[\s\S]*?([\d.eE,+\s-]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"-?[\d.eE+-]+", match.group(1))
            cpen_srk_over_r = [float(value) for value in raw_values][2:]
            cpen = [value * self.R * 1000 for value in cpen_srk_over_r]
            logger.info("CpenSRKExtractor: успешно извлечены cpen")
            return {'cpen': cpen}
        else:
            logger.warning("Тег <Cpen SRK> не найден в файле.")
            return {'cpen': []}

class CpenPRExtractor(CHCExtractor):
    R = 0.082057  # Газовая постоянная в L·atm/(K·mol)

    def extract(self) -> Dict[str, List[float]]:
        match = re.search(r"<Cpen PR>[\s\S]*?([\d.eE,+\s-]+)\n(?=<|$)", self.content)
        if match:
            raw_values = re.findall(r"-?[\d.eE+-]+", match.group(1))
            cpen_pr_over_r = [float(value) for value in raw_values][2:]
            cpen_pr = [value * self.R * 1000 for value in cpen_pr_over_r]
            logger.info("CpenPRExtractor: успешно извлечены cpen")
            return {'cpen': cpen_pr}
        else:
            logger.warning("Тег <Cpen PR> не найден в файле.")
            return {'cpen': []}

# === Парсеры PVTSim для SRK и PR ===
class PVTSimParserSRK(Parser):
    def load_data(self, file_path: str) -> Optional[Dict]:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                content = file.read()
        except FileNotFoundError:
            logger.error(f"{self.__class__.__name__}: файл {file_path} не найден.")
            return None
        except Exception as e:
            logger.error(f"{self.__class__.__name__}: ошибка при чтении файла {file_path}: {e}")
            return None

        # Инициализируем экстракторы
        extractors = [
            ShortCompNameExtractor(content),
            CompAmountExtractor(content),
            MolecularWeightsExtractor(content),
            CriticalPressureExtractor(content),
            CriticalTemperatureExtractor(content),
            CriticalVolumeExtractor(content),
            AcentricFactorExtractor(content),
            BoilingTemperatureExtractor(content),
            LiquidDensityExtractor(content),
            CpenSRKExtractor(content)
        ]

        # Собираем данные
        data = {}
        for extractor in extractors:
            extracted_data = extractor.extract()
            if isinstance(extracted_data, dict):
                data.update(extracted_data)
            else:
                logger.warning(f"{extractor.__class__.__name__}: extract метод вернул некорректный тип данных.")

        return data

class PVTSimParserPR(Parser):
    def load_data(self, file_path: str) -> Optional[Dict]:
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                content = file.read()
        except FileNotFoundError:
            logger.error(f"{self.__class__.__name__}: файл {file_path} не найден.")
            return None
        except Exception as e:
            logger.error(f"{self.__class__.__name__}: ошибка при чтении файла {file_path}: {e}")
            return None

        # Инициализируем экстракторы
        extractors = [
            ShortCompNameExtractor(content),
            CompAmountExtractor(content),
            MolecularWeightsExtractor(content),
            CriticalPressureExtractor(content),
            CriticalTemperatureExtractor(content),
            CriticalVolumeExtractor(content),
            AcentricFactorExtractor(content),
            BoilingTemperatureExtractor(content),
            LiquidDensityExtractor(content),
            CpenPRExtractor(content)
        ]

        # Собираем данные
        data = {}
        for extractor in extractors:
            extracted_data = extractor.extract()
            if isinstance(extracted_data, dict):
                data.update(extracted_data)
            else:
                logger.warning(f"{extractor.__class__.__name__}: extract метод вернул некорректный тип данных.")

        return data

# === Маппер для данных ===
class FluidMapper:
    @staticmethod
    def map_to_fluid_dto(data: Dict, p: float, t: float, bips: Optional[pd.DataFrame] = None) -> FluidDTO:
        component_names = data.get("component_name", [])
        z = data.get("z", [])
        mass = data.get("mass", [])
        Pkr = data.get("Pkr", [])
        Tkr = data.get("Tkr", [])
        Vkr = data.get("Vkr", [])
        w = data.get("w", [])
        cpen = data.get("cpen", [])
        T_boil = data.get("T_boil", [])
        density_liq_phase = data.get("density_liq_phase", [])

        num_components = len(component_names)
        # Проверка согласованности данных
        for key, value in [("z", z), ("mass", mass), ("Pkr", Pkr), ("Tkr", Tkr),
                           ("Vkr", Vkr), ("w", w), ("cpen", cpen),
                           ("T_boil", T_boil), ("density_liq_phase", density_liq_phase)]:
            if len(value) != num_components:
                raise ValueError(f"Длина данных для {key} ({len(value)}) не совпадает с количеством компонентов ({num_components}).")

        components = [
            FluidComponentDTO(
                component_name=component_names[i],
                z=float(z[i]),
                mass=float(mass[i]),
                Pkr=float(Pkr[i]),
                Tkr=float(Tkr[i]),
                Vkr=float(Vkr[i]),
                w=float(w[i]),
                cpen=float(cpen[i]),
                T_boil=float(T_boil[i]),
                density_liq_phase=float(density_liq_phase[i]),
            )
            for i in range(num_components)
        ]
        logger.info("FluidMapper: успешно преобразованы данные в FluidDTO")
        return FluidDTO(components=components, BIPs=bips, p=p, t=t)

# === Универсальный обработчик данных ===
class FluidDataProcessor:
    def __init__(self, parser: Parser):
        self.parser = parser

    def process(self, file_path: str, p: float, t: float, bips_path: Optional[str] = None) -> FluidDTO:
        # Загрузка данных
        data = self.parser.load_data(file_path)
        if not data:
            raise ValueError("Не удалось загрузить данные.")

        # Загрузка BIPs, если указан путь
        bips = None
        if bips_path:
            if isinstance(self.parser, ExcelParser):
                bips = data.get('BIPs')
                if bips is None:
                    raise ValueError("Не удалось загрузить данные BIPs из Excel.")
            else:
                bips_parser = ExcelBIPSParser()
                bips = bips_parser.load_data(bips_path)
                if bips is None:
                    raise ValueError("Не удалось загрузить данные BIPs.")

        # Маппинг данных в FluidDTO
        return FluidMapper.map_to_fluid_dto(data, p, t, bips)

# === Основной блок ===
if __name__ == "__main__":
    # === Обработка Excel данных ===
    excel_parser = ExcelParser()
    processor_excel = FluidDataProcessor(parser=excel_parser)
    try:
        fluid_excel = processor_excel.process(
            file_path="data/input_data.xlsx",
            p=15,
            t=473.15,
            bips_path=None  # Укажите путь к BIPs, если необходимо
        )
        calculated_SRK_Peneloux_excel = SRK_peneloux_Flash(fluid_excel)
        srk_peneloux_excel = calculated_SRK_Peneloux_excel.calculate()

        print("Результаты из Excel данных:")
        for each_elem in srk_peneloux_excel:
            print(f"{each_elem}")
    except ValueError as ve:
        logger.error(f"Ошибка при обработке Excel данных: {ve}")

    # === Обработка CHC данных для SRK ===
    chc_file_path = "Mixture1 (1).CHC"
    pvtsim_srk_parser = PVTSimParserSRK()
    processor_srk = FluidDataProcessor(parser=pvtsim_srk_parser)
    try:
        fluid_srk = processor_srk.process(
            file_path=chc_file_path,
            p=15,
            t=473.15,
            bips_path=None  # Укажите путь к BIPs, если необходимо
        )
        calculated_SRK_Peneloux_srk = SRK_peneloux_Flash(fluid_srk)
        srk_peneloux_srk = calculated_SRK_Peneloux_srk.calculate()

        print("\nРезультаты из CHC данных (SRK):")
        for each_elem in srk_peneloux_srk:
            print(f"{each_elem}")
    except ValueError as ve:
        logger.error(f"Ошибка при обработке CHC данных (SRK): {ve}")

    # === Обработка CHC данных для PR ===
    pvtsim_pr_parser = PVTSimParserPR()
    processor_pr = FluidDataProcessor(parser=pvtsim_pr_parser)
    try:
        fluid_pr = processor_pr.process(
            file_path=chc_file_path,
            p=15,
            t=473.15,
            bips_path=None  # Укажите путь к BIPs, если необходимо
        )
        calculated_SRK_Peneloux_pr = SRK_peneloux_Flash(fluid_pr)
        srk_peneloux_pr = calculated_SRK_Peneloux_pr.calculate()

        print("\nРезультаты из CHC данных (PR):")
        for each_elem in srk_peneloux_pr:
            print(f"{each_elem}")
    except ValueError as ve:
        logger.error(f"Ошибка при обработке CHC данных (PR): {ve}")
